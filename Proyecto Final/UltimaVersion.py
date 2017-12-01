from openpyxl import load_workbook

wb = load_workbook(filename='template base de datos.xlsx')

wb.get_sheet_names()
sheet1 = wb.get_sheet_by_name('Venezuela')
sheet2 = wb.get_sheet_by_name('Peru')
sheet3 = wb.get_sheet_by_name('Viajes')
sheet4 = wb.get_sheet_by_name('Reporte final')

def imprimir_pais():
    for x in paises:
        print(x, "->", paises[x])

def solicitar_servicios(n_servicios_pais):
    lista1 = []
    for i in range(n_servicios_pais):
        servicios_pais = input("Ingrese servicio #" + str(i+1) + ": ")
        lista1.append(servicios_pais)
    return lista1

def solicitar_meses(n_meses_pais):
    lista2 = []
    for i in range(n_meses_pais):
        meses_pais = input("Ingrese mes #" + str(i+1) + ": ")
        lista2.append(meses_pais)
    return lista2

def buscar_gastos(hoja, inicio_i, fin_i, paso, inicio_j, fin_j):
    lista3 = []
    for item in lista1:
        for i in range(inicio_i, fin_i, paso):
            if item == hoja.cell(row=1, column = i).internal_value:
                for mes in lista2:
                    for j in range(inicio_j, fin_j, paso):
                        if mes == hoja.cell(row=j, column=2).internal_value:
                            lista3.append(hoja.cell(row=j, column = i).internal_value)
    return lista3

def imprimir_gasto_opc1(n_servicios_pais, n_meses_pais):
    h = 0
    for i in range(n_servicios_pais):
        for j in range(n_meses_pais):
            print("Para el servicio #" + str(i+1) + " en el mes #" + str(j+1) + ", el gasto es: ", lista3[h])
            h += 1

def imprimir_meses(hoja):
    for i in range(2, 14, 1):
        print(hoja.cell(row=i, column=2).internal_value)

def imprimir_gasto_opc2(hoja, columna):
    for i in range(2, 14, 1):
        if mes_seleccionado == hoja.cell(row=i, column=2).internal_value:
            print(hoja.cell(row=i, column=columna).value)

def imprimir_servicios(hoja, fin):
    for i in range(3, fin, 1):
        print(hoja.cell(row=1, column=i).internal_value)

def imprimir_gasto_opc3(hoja, fin):
    for i in range(3, fin, 1):
        if servicio_seleccionado == hoja.cell(row=1, column=i).internal_value:
            print(hoja.cell(row=14, column=i).value)

def imprimir_reporte():
    for row in sheet4.iter_rows():
        for i, cell in enumerate(row):
            value = " "
            if cell.internal_value is not None:
                value = cell.internal_value
            if type(str()) == type(value):
                print("{:^15s} ".format(value), end="")
            elif type(value) in [type(int()), type(float())] and i != 0:
                print("{:12,.2f} ".format(value), end="")
            elif type(value) in [type(int()), type(float())] and i == 0:
                print("{:^15s} ".format(str(value)), end="")
            else:
                print(value, type(value))
        print(" ")


print("¡Bienvenido Carlos!. ¿En qué te puedo ayudar?")

print()

opcion_seleccionada = -1
while opcion_seleccionada != 0:
    opciones_de_entrada = (
    "1 -> Consultar el gasto asociado a uno o varios servicios correspondientes a uno o varios meses",
    "2 -> Consultar el gasto total de todos los servicios en un mes particular",
    "3 -> Consultar el gasto total anual de un servicio particular",
    "4 -> Consultar reporte del total a pagar",
    "0 -> Salir")
    for i in range(len(opciones_de_entrada)):
        print(opciones_de_entrada[i])

    print()

    opcion_seleccionada = int(input("Introduce el número de tu opción seleccionada: "))

    paises = {"Venezuela":1, "Perú":2}

    if opcion_seleccionada == 1:
        print()
        imprimir_pais()
        print()
        pais = int(input("Indique el número de la opción que desea: "))
        print()
        if pais == 1:
            n_servicios_pais = int(input("Ingrese cantidad de servicios a consultar: "))
            print()
            for i in range(3, 18, 1):
                print(sheet1.cell(row=1, column=i).value)
            print()
            lista1 = solicitar_servicios(n_servicios_pais)
            print()
            n_meses_pais = int(input("Ingrese cantidad de meses a consultar: "))
            lista2 = solicitar_meses(n_meses_pais)
            lista3 = buscar_gastos(sheet1, 3, 18, 1, 2, 14)
            print()
            imprimir_gasto_opc1(n_servicios_pais, n_meses_pais)
        elif pais == 2:
            n_servicios_pais = int(input("Ingrese cantidad de servicios a consultar: "))
            print()
            for i in range(3, 13, 1):
                print(sheet2.cell(row=1, column=i).value)
            print()
            lista1 = solicitar_servicios(n_servicios_pais)
            print()
            n_meses_pais = int(input("Ingrese cantidad de meses a consultar: "))
            lista2 = solicitar_meses(n_meses_pais)
            lista3 = buscar_gastos(sheet2, 3, 13, 1, 2, 14)
            print()
            imprimir_gasto_opc1(n_servicios_pais, n_meses_pais)
        else:
            print("Dato inválido. Vuelva a intentar")

    elif opcion_seleccionada == 2:
        print()
        imprimir_pais()
        print()
        pais = int(input("Indique el número de la opción que desea: "))
        print()
        if pais == 1:
            meses = imprimir_meses(sheet1)
            print()
            mes_seleccionado = input("Indique el mes: ")
            print()
            gasto = imprimir_gasto_opc2(sheet1, 18)
        elif pais == 2:
            meses = imprimir_meses(sheet2)
            print()
            mes_seleccionado = input("Indique el mes: ")
            print()
            gasto = imprimir_gasto_opc2(sheet2, 13)
        else:
            print("Dato inválido. Vuelva a intentar")

    elif opcion_seleccionada == 3:
        print()
        imprimir_pais()
        print()
        pais = int(input("Indique el número de la opción que desea: "))
        print()
        if pais == 1:
            servicios = imprimir_servicios(sheet1, 18)
            print()
            servicio_seleccionado = input("Indique el servicio: ")
            print()
            gasto = imprimir_gasto_opc3(sheet1, 18)
        elif pais == 2:
            servicios = imprimir_servicios(sheet2, 13)
            print()
            servicio_seleccionado = input("Indique el servicio: ")
            print()
            gasto = imprimir_gasto_opc3(sheet2, 13)
        else:
            print("Dato inválido. Vuelva a intentar")

    elif opcion_seleccionada == 4:
        reporte = imprimir_reporte()
    elif opcion_seleccionada == 0:
        break
    else:
        print("Dato inválido. Vuelva a intentar")
    input("Presione enter para continuar...")
    print()
